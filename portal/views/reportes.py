# ==============================================================================
#  🛡️ SISTEMA DIGITAL DE PROTECCIÓN CIVIL Y BOMBEROS
#  Copyright (c) 2026 Josué Jaziel Delgado Burela. Todos los derechos reservados.
#  Desarrollado y Diseñado por: Josué Jaziel Delgado Burela
#  Contacto y Soporte: jburela1@gmal.com
# ==============================================================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from portal.models import ReporteRiesgo, Ciudadano, PersonalAdministrativo, HistorialReporte, Localidad
from portal.utils.whatsapp_utils import enviar_alerta_grupo, enviar_actualizacion_grupo
import logging

logger = logging.getLogger(__name__)

from django.db import models

def crear_reporte(request):
    ciudadano = None
    nombre_prellenado = ""
    tel_prellenado = ""
    ciudadano_id = request.session.get('ciudadano_id')
    curp_sesion = request.session.get('ciudadano_curp')
    if ciudadano_id:
        ciudadano = Ciudadano.objects.filter(id=ciudadano_id).first()
    elif curp_sesion:
        ciudadano = Ciudadano.objects.filter(models.Q(curp=curp_sesion) | models.Q(correo=curp_sesion)).first()
        
    if ciudadano:
        nombre_prellenado = f"{ciudadano.nombre} {ciudadano.primer_apellido} {ciudadano.segundo_apellido or ''}".strip()
        tel_prellenado = ciudadano.telefono

    if request.method == 'POST':
        nombre = request.POST.get('nombre_ciudadano', '').strip()
        telefono = request.POST.get('telefono_ciudadano', '').strip()
        tipo = request.POST.get('tipo_servicio', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        colonia = request.POST.get('colonia', '').strip()
        localidad = request.POST.get('localidad', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        foto = request.FILES.get('evidencia_foto')
        prioridad = request.POST.get('prioridad', 'MEDIA')
        latitud_val = request.POST.get('latitud', '').strip()
        longitud_val = request.POST.get('longitud', '').strip()

        try:
            lat_float = float(latitud_val) if latitud_val else None
            lng_float = float(longitud_val) if longitud_val else None
        except ValueError:
            lat_float = None
            lng_float = None

        if nombre and telefono and tipo and direccion and colonia and localidad and descripcion:
            reporte = ReporteRiesgo.objects.create(
                ciudadano=ciudadano,
                nombre_ciudadano=nombre,
                telefono_ciudadano=telefono,
                tipo_servicio=tipo,
                direccion=direccion,
                colonia=colonia,
                localidad=localidad,
                latitud=lat_float,
                longitud=lng_float,
                descripcion=descripcion,
                evidencia_foto=foto,
                prioridad=prioridad
            )
            
            # Enviar alerta al grupo de WhatsApp (pausado temporalmente por preferencia de recursos)
            DESHABILITAR_ALERTAS_WHATSAPP = True
            if not DESHABILITAR_ALERTAS_WHATSAPP:
                try:
                    enviar_alerta_grupo(reporte)
                except Exception as e:
                    logger.error(f"Error al enviar alerta WhatsApp para {reporte.numero_reporte}: {e}")
            
            messages.success(request, f"¡Reporte enviado con éxito! Se ha generado tu folio de seguimiento: {reporte.numero_reporte}.")
            return redirect(f'/crear-reporte/?success_folio={reporte.numero_reporte}')
        else:
            messages.error(request, "Por favor, completa todos los campos obligatorios.")
            
    localidades = Localidad.objects.all().order_by('nombre')
    return render(request, 'portal/crear_reporte.html', {
        'nombre_prellenado': nombre_prellenado,
        'tel_prellenado': tel_prellenado,
        'localidades': localidades
    })

@login_required(login_url='login_unificado')
def asignar_reporte(request, reporte_id):
    reporte = get_object_or_404(ReporteRiesgo, id=reporte_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        unidad = request.POST.get('unidad_acudira', '').strip()
        dependencia = request.POST.get('dependencia_conocimiento', '').strip()
        tiempo = request.POST.get('tiempo_atencion', '').strip()
        responsables_ids = request.POST.getlist('responsables')
        
        # Determinar estatus de forma automática basado en la acción y campos ingresados
        if action == 'finalizar':
            estatus = 'RESUELTO'
            if not unidad:
                unidad = reporte.unidad_acudira or "Unidad de Respuesta PC"
            if not tiempo:
                tiempo = reporte.tiempo_atencion or "Completado"
        else:
            # Si intentaron ingresar parcialmente datos de despacho, requerir ambos
            if (unidad and not tiempo) or (tiempo and not unidad):
                messages.error(request, "Error: La Unidad Despachada y el Tiempo Estimado de Atención son obligatorios para cambiar el estatus a 'En Camino'.")
                return redirect(f'/panel/?seccion=reportes&ver_reporte={reporte.id}')
                
            if unidad and tiempo:
                estatus = 'EN_PROCESO'
            else:
                if reporte.estatus == 'EN_PROCESO':
                    estatus = 'EN_PROCESO'
                elif reporte.estatus == 'RESUELTO':
                    estatus = 'RESUELTO'
                else:
                    estatus = 'LEIDO'

        # Detectar si hubo cambios importantes en los valores para agregarlos automáticamente a la bitácora
        cambios = []
        if reporte.estatus != estatus:
            cambios.append(f"Cambió estatus de '{reporte.get_estatus_display()}' a '{dict(ReporteRiesgo.ESTATUS_CHOICES).get(estatus)}'")
        if reporte.unidad_acudira != unidad and unidad:
            cambios.append(f"Asignó unidad: '{unidad}'")
            
        # Actualizar datos principales
        if estatus == 'RESUELTO' and reporte.estatus != 'RESUELTO':
            from django.utils import timezone
            reporte.fecha_resolucion = timezone.now()
        elif estatus != 'RESUELTO':
            reporte.fecha_resolucion = None
            
        reporte.estatus = estatus
        reporte.unidad_acudira = unidad
        reporte.dependencia_conocimiento = dependencia
        reporte.tiempo_atencion = tiempo
        
        # Guardar relaciones ManyToMany ("Quienes responderán")
        reporte.responsables.clear()
        if responsables_ids:
            responsables = PersonalAdministrativo.objects.filter(id__in=responsables_ids)
            reporte.responsables.add(*responsables)
            
        reporte.save()
        
        # Registrar cambios en la bitácora automáticamente si existen
        if cambios:
            HistorialReporte.objects.create(
                reporte=reporte,
                creado_por=request.user,
                comentario=f"[Control Operativo] {', '.join(cambios)}"
            )
            
            # Enviar actualización al grupo de WhatsApp (fire-and-forget)
            try:
                cambios_texto = "\n".join([f"• {c}" for c in cambios])
                enviar_actualizacion_grupo(reporte, cambios_texto)
            except Exception as e:
                logger.error(f"Error al enviar actualización WhatsApp para {reporte.numero_reporte}: {e}")
            
        messages.success(request, f"Reporte {reporte.numero_reporte} actualizado correctamente.")
        
    return redirect(f'/panel/?seccion=reportes&ver_reporte={reporte.id}')

@login_required(login_url='login_unificado')
def agregar_bitacora(request, reporte_id):
    reporte = get_object_or_404(ReporteRiesgo, id=reporte_id)
    
    if request.method == 'POST':
        comentario = request.POST.get('comentario', '').strip()
        if comentario:
            HistorialReporte.objects.create(
                reporte=reporte,
                creado_por=request.user,
                comentario=comentario
            )
            messages.success(request, "Bitácora actualizada correctamente.")
        else:
            messages.error(request, "No puedes agregar una entrada vacía a la bitácora.")
            
    return redirect(f'/panel/?seccion=reportes&ver_reporte={reporte.id}')

def consultar_reporte(request):
    folio = request.GET.get('folio', '').strip().upper()
    reporte = None
    historial = None
    chat_mensajes = None
    buscado = False
    
    if folio:
        buscado = True
        reporte = ReporteRiesgo.objects.filter(numero_reporte=folio).first()
        if reporte:
            historial = reporte.historial.all().order_by('-fecha_registro')
            chat_mensajes = reporte.mensajes_chat.all().order_by('fecha_envio')
            
    # Si viene por AJAX (para actualizaciones en tiempo real)
    if request.GET.get('ajax') == '1':
        from django.http import JsonResponse
        if reporte:
            # Calcular una huella digital (hash) del estado actual de este reporte
            import hashlib
            raw_state = f"{reporte.estatus}|{reporte.unidad_acudira or ''}|{reporte.tiempo_atencion or ''}|{reporte.historial.count()}|{reporte.mensajes_chat.count()}"
            current_hash = hashlib.md5(raw_state.encode('utf-8')).hexdigest()
            
            # Si el cliente ya tiene esta misma versión, responder inmediatamente que no hay cambios
            client_hash = request.GET.get('version', '')
            if client_hash == current_hash:
                return JsonResponse({
                    'encontrado': True,
                    'modificado': False,
                    'version': current_hash
                })
                
            historial_data = [
                {
                    'comentario': h.comentario,
                    'fecha': h.fecha_registro.strftime('%d/%m/%Y %H:%M')
                } for h in (historial or [])
            ]
            # Agregar la entrada inicial estática del reporte lanzado
            historial_data.append({
                'comentario': f"Reporte lanzado e ingresado en la central municipal. Folio generado: {reporte.numero_reporte}.",
                'fecha': reporte.fecha_reporte.strftime('%d/%m/%Y %H:%M')
            })
            
            chat_data = [
                {
                    'remitente': m.remitente_admin.username if m.remitente_admin else (m.remitente_ciudadano.nombre if m.remitente_ciudadano else "Ciudadano"),
                    'is_admin': m.remitente_admin is not None,
                    'mensaje': m.mensaje,
                    'fecha': m.fecha_envio.strftime('%d/%m/%Y %H:%M')
                } for m in (chat_mensajes or [])
            ]
            
            return JsonResponse({
                'encontrado': True,
                'modificado': True,
                'version': current_hash,
                'numero_reporte': reporte.numero_reporte,
                'estatus': reporte.estatus,
                'estatus_display': reporte.get_estatus_display(),
                'tipo_servicio': reporte.get_tipo_servicio_display(),
                'fecha_reporte': reporte.fecha_reporte.strftime('%d/%m/%Y %H:%M'),
                'direccion': reporte.direccion,
                'colonia': reporte.colonia,
                'localidad': reporte.localidad,
                'descripcion': reporte.descripcion,
                'unidad_acudira': reporte.unidad_acudira or 'Por asignar brigada',
                'tiempo_atencion': reporte.tiempo_atencion or 'En evaluación operativa',
                'prioridad_display': reporte.get_prioridad_display(),
                'tiempo_resolucion_calculado': reporte.tiempo_resolucion_calculado,
                'historial': historial_data,
                'chat_mensajes': chat_data
            })
        else:
            return JsonResponse({'encontrado': False})

    return render(request, 'portal/consultar_reporte.html', {
        'folio': folio,
        'reporte': reporte,
        'historial': historial,
        'chat_mensajes': chat_mensajes,
        'buscado': buscado
    })


# ==============================================================================
#  📱 HERRAMIENTA WEB DE CLASIFICACIÓN E IMPORTACIÓN DE CHATS DE WHATSAPP
# ==============================================================================

import re
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from portal.views.vehiculos import requiere_operador_aprobado

CATEGORIAS_PATTERNS = {
    'FUEGO': [r'incendio', r'fuego', r'pastizal', r'humo', r'quema', r'basura', r'llanta', r'se quema'],
    'CHOQUE': [r'choque', r'accidente', r'volcadura', r'atropellad', r'derrapad', r'moto', r'auto', r'vehiculo', r'carro'],
    'GAS': [r'gas', r'fuga', r'tanque', r'cilindro', r'olor a gas', r'oler'],
    'ABEJAS': [r'abeja', r'enjambre', r'avispa', r'picadura'],
    'ARBOL': [r'arbol', r'árbol', r'rama', r'caido', r'caído', r'viento'],
    'CABLE': [r'cable', r'poste', r'transformador', r'luz', r'corto', r'chispas'],
    'AGUA': [r'inundac', r'agua', r'rio', r'río', r'anegad', r'inundado', r'drenaje', r'canal'],
    'OTRO': []
}

LOCALIDADES_MEDELLIN = [
    'PUENTE MORENO', 'ARBOLEDAS SAN RAMÓN', 'LAGOS DE PUENTE MORENO', 'EL TEJAR',
    'MEDELLÍN', 'PLAYA DE VACAS', 'PASO DEL TORO', 'LOS ROBLES', 'DOS BOCAS',
    'RANCHO DEL PADRE', 'PASO COLORADO', 'LA JOYA', 'PASEO CAMPESTRE', 'HERÓN PROAL',
    'MARCOS VÉLEZ', 'EMILIANO ZAPATA', 'CLARA CÓRDOBA'
]

@requiere_operador_aprobado
def importar_chat_whatsapp_web(request):
    """
    Herramienta Web para pegar el texto del chat de WhatsApp o subir un archivo .txt,
    parsear reportes del mes indicado (ej. Agosto), clasificarlos y exportar a Excel o BD.
    """
    reportes_filtrados = []
    texto_raw = ""
    mes_target = int(request.POST.get('mes', 8))
    year_target = int(request.POST.get('year', 2026))
    action = request.POST.get('action', '')

    if request.method == 'POST':
        if 'archivo_txt' in request.FILES:
            texto_raw = request.FILES['archivo_txt'].read().decode('utf-8', errors='ignore')
        else:
            texto_raw = request.POST.get('texto_chat', '')

        if texto_raw:
            lines = texto_raw.splitlines()
            pattern_ts = re.compile(r'^(?:\[)?(\d{1,2}/\d{1,2}/\d{2,4})[,\s]+(\d{1,2}:\d{2}(?::\d{2})?(?:\s*[ap]\.?\s*m\.?)?)(?:\])?\s*[-–]?\s*(.*?):\s*(.*)$', re.IGNORECASE)

            reportes_extraidos = []
            msg_buffer = None

            for line in lines:
                line_str = line.strip()
                match = pattern_ts.match(line_str)

                if match:
                    if msg_buffer:
                        reportes_extraidos.append(msg_buffer)
                        msg_buffer = None

                    fecha_raw, hora_raw, remitente, texto = match.groups()

                    try:
                        parts = fecha_raw.split('/')
                        d = int(parts[0])
                        m = int(parts[1])
                        y = int(parts[2])
                        if y < 100: y += 2000
                        dt = datetime.date(y, m, d)
                    except Exception:
                        continue

                    if dt.month == mes_target and dt.year == year_target:
                        msg_buffer = {
                            'fecha_str': dt.strftime('%d/%m/%Y'),
                            'hora': hora_raw.strip(),
                            'remitente': remitente.strip(),
                            'texto': texto.strip()
                        }
                else:
                    if msg_buffer:
                        msg_buffer['texto'] += f" {line_str}"

            if msg_buffer:
                reportes_extraidos.append(msg_buffer)

            for r in reportes_extraidos:
                txt_lower = r['texto'].lower()
                if any(sys_kw in txt_lower for sys_kw in ['cambió', 'añadió', 'eliminó', 'salio', 'salió', 'cifrado', 'código de seguridad']):
                    continue
                if len(r['texto']) < 5:
                    continue

                tipo_detectado = 'OTRO'
                for cat, kw_list in CATEGORIAS_PATTERNS.items():
                    if any(re.search(kw, txt_lower) for kw in kw_list):
                        tipo_detectado = cat
                        break
                r['tipo_servicio'] = tipo_detectado

                loc_detectada = 'MEDELLÍN'
                for loc in LOCALIDADES_MEDELLIN:
                    if loc.lower() in txt_lower:
                        loc_detectada = loc
                        break
                r['localidad'] = loc_detectada

                reportes_filtrados.append(r)

            # SI SE SOLICITA EXPORTAR A EXCEL DIRECTAMENTE
            if action == 'descargar_excel' and reportes_filtrados:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Reportes Mes {mes_target}"

                header_fill = PatternFill(start_color="5A123E", end_color="5A123E", fill_type="solid")
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                title_font = Font(name="Calibri", size=14, bold=True, color="5A123E")
                border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                                top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

                ws.merge_cells("A1:G1")
                ws["A1"] = f"DIRECCIÓN DE PROTECCIÓN CIVIL — REPORTES EXTRAÍDOS DE WHATSAPP ({mes_target}/{year_target})"
                ws["A1"].font = title_font

                headers = ["#", "Fecha", "Hora", "Remitente", "Categoría Incidente", "Colonia / Localidad", "Descripción del Reporte"]
                ws.append([])
                ws.append(headers)

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for idx, item in enumerate(reportes_filtrados, 1):
                    row = [idx, item['fecha_str'], item['hora'], item['remitente'], item['tipo_servicio'], item['localidad'], item['texto']]
                    ws.append(row)
                    r_num = ws.max_row
                    for c_i in range(1, 8):
                        c = ws.cell(row=r_num, column=c_i)
                        c.border = border
                        if c_i in [1, 2, 3, 5, 6]:
                            c.alignment = Alignment(horizontal="center")

                response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                response['Content-Disposition'] = f'attachment; filename="Reportes_WhatsApp_Mes_{mes_target}_{year_target}.xlsx"'
                wb.save(response)
                return response

            # SI SE SOLICITA IMPORTAR A BASE DE DATOS
            elif action == 'importar_db' and reportes_filtrados:
                creados = 0
                for item in reportes_filtrados:
                    count = ReporteRiesgo.objects.count() + 1
                    folio = f"REP-{year_target}-WA-{count:04d}"

                    ReporteRiesgo.objects.create(
                        numero_reporte=folio,
                        nombre_ciudadano=item['remitente'][:150],
                        telefono_ciudadano='2290000000',
                        tipo_servicio=item['tipo_servicio'],
                        descripcion=item['texto'],
                        localidad=item['localidad'],
                        ubicacion_direccion=f"Reportado vía WhatsApp en {item['localidad']}",
                        estatus='RESUELTO'
                    )
                    creados += 1
                messages.success(request, f"¡Éxito! Se importaron {creados} reportes a la base de datos de Protección Civil.")
                return redirect('flotilla_vehiculos_hub')

    return render(request, 'portal/reportes_importar_chat.html', {
        'reportes_filtrados': reportes_filtrados,
        'texto_raw': texto_raw,
        'mes_target': mes_target,
        'year_target': year_target,
        'operador_actual': request.operador_actual
    })
