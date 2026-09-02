# ==============================================================================
#  🎓 MÓDULO DE CAPACITACIONES Y EMISIÓN AUTOMATIZADA DE CONSTANCIAS
#  Copyright (c) 2026 Josué Jaziel Delgado Burela. Todos los derechos reservados.
# ==============================================================================

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from portal.models import CursoCapacitacion, InscripcionCapacitacion
from portal.views.vehiculos import requiere_operador_aprobado


def registro_capacitacion_publico(request):
    """
    Formulario Público de Registro a Cursos de Capacitación impartidos por Protección Civil.
    Solo muestra cursos activos y que NO han sido finalizados / cerrados.
    """
    cursos_activos = CursoCapacitacion.objects.filter(activo=True, finalizado=False).order_by('-fecha_inicio')
    
    if request.method == 'POST':
        curso_id = request.POST.get('curso_id')
        nombre = request.POST.get('nombre_completo', '').strip()
        curp = request.POST.get('curp', '').strip()
        correo = request.POST.get('correo', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        empresa = request.POST.get('empresa_institucion', '').strip()
        
        curso = get_object_or_404(CursoCapacitacion, id=curso_id, activo=True, finalizado=False)
        
        if not nombre:
            messages.error(request, "Por favor ingresa el nombre completo del participante.")
            return render(request, 'portal/capacitacion_registro_publico.html', {
                'cursos_activos': cursos_activos
            })
            
        # Verificar si ya está inscrito en el mismo curso con el mismo correo (si se proporcionó)
        if correo and InscripcionCapacitacion.objects.filter(curso=curso, correo=correo).exists():
            inscripcion_existente = InscripcionCapacitacion.objects.filter(curso=curso, correo=correo).first()
            messages.warning(request, f"Ya existe una inscripción registrada para este curso a nombre de {inscripcion_existente.nombre_completo}.")
            return render(request, 'portal/capacitacion_registro_publico.html', {
                'cursos_activos': cursos_activos,
                'inscripcion_exitosa': inscripcion_existente
            })
            
        # Crear la nueva inscripción
        inscripcion = InscripcionCapacitacion.objects.create(
            curso=curso,
            nombre_completo=nombre,
            curp=curp or None,
            correo=correo or None,
            telefono=telefono or None,
            empresa_institucion=empresa or None
        )
        
        messages.success(request, f"¡Inscripción registrada con éxito! Tu folio de constancia asignado es {inscripcion.folio_constancia}.")
        return render(request, 'portal/capacitacion_registro_publico.html', {
            'cursos_activos': cursos_activos,
            'inscripcion_exitosa': inscripcion
        })

    return render(request, 'portal/capacitacion_registro_publico.html', {
        'cursos_activos': cursos_activos
    })


@requiere_operador_aprobado
def admin_capacitaciones_dashboard(request):
    """
    Panel de Administración de Capacitaciones para el personal operativo y directivo.
    Distingue entre Cursos Activos y Cursos Pasados / Finalizados.
    """
    cursos_todos = CursoCapacitacion.objects.all().order_by('-fecha_inicio')
    cursos_activos = cursos_todos.filter(finalizado=False)
    cursos_pasados = cursos_todos.filter(finalizado=True)
    
    curso_seleccionado_id = request.GET.get('curso_id')
    
    if curso_seleccionado_id:
        curso_actual = get_object_or_404(CursoCapacitacion, id=curso_seleccionado_id)
        inscripciones = InscripcionCapacitacion.objects.filter(curso=curso_actual).order_by('-fecha_registro')
    else:
        # Priorizar mostrar el primer curso activo, si no hay, el primero general
        curso_actual = cursos_activos.first() or cursos_todos.first()
        if curso_actual:
            inscripciones = InscripcionCapacitacion.objects.filter(curso=curso_actual).order_by('-fecha_registro')
        else:
            inscripciones = InscripcionCapacitacion.objects.none()
            
    return render(request, 'portal/capacitacion_admin_dashboard.html', {
        'cursos': cursos_todos,
        'cursos_activos': cursos_activos,
        'cursos_pasados': cursos_pasados,
        'curso_actual': curso_actual,
        'inscripciones': inscripciones,
        'operador_actual': request.operador_actual
    })


@requiere_operador_aprobado
def toggle_finalizar_curso_admin(request, curso_id):
    """
    Alterna el estatus de un taller entre Activo y Finalizado / Pasado.
    Al finalizar un curso se cierran automáticamente las inscripciones públicas.
    """
    curso = get_object_or_404(CursoCapacitacion, id=curso_id)
    curso.finalizado = not curso.finalizado
    if curso.finalizado:
        curso.activo = False
        msg = f"El taller '{curso.titulo}' ha sido marcado como FINALIZADO. Se cerraron las inscripciones y pasó a Cursos Pasados."
    else:
        curso.activo = True
        msg = f"El taller '{curso.titulo}' ha sido REABIERTO y vuelve a admitir inscripciones."
        
    curso.save()
    messages.success(request, msg)
    return redirect(f'/capacitaciones/admin/?curso_id={curso.id}')


@requiere_operador_aprobado
def crear_curso_admin(request):
    """
    Crea un nuevo curso de capacitación desde el panel de administración.
    """
    if request.method == 'POST':
        titulo = request.POST.get('titulo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        duracion_horas = request.POST.get('duracion_horas', 8)
        fecha_inicio = request.POST.get('fecha_inicio')
        horario = request.POST.get('horario', '09:00 a 14:00 hrs').strip()
        sede = request.POST.get('sede_ubicacion', 'Estación Central de Bomberos El Tejar').strip()
        cupo = request.POST.get('cupo_maximo', 50)
        
        if titulo and fecha_inicio:
            curso = CursoCapacitacion.objects.create(
                titulo=titulo,
                descripcion=descripcion,
                duracion_horas=duracion_horas,
                fecha_inicio=fecha_inicio,
                horario=horario,
                sede_ubicacion=sede,
                cupo_maximo=cupo,
                activo=True
            )
            messages.success(request, f"Curso '{curso.titulo}' creado exitosamente.")
        else:
            messages.error(request, "Por favor ingresa un título y fecha de inicio válidos.")
            
    return redirect('admin_capacitaciones_dashboard')


@requiere_operador_aprobado
def marcar_asistencia_capacitacion(request, inscripcion_id):
    """
    Alterna el estado de asistencia y aprobación de un participante.
    """
    inscripcion = get_object_or_404(InscripcionCapacitacion, id=inscripcion_id)
    
    # Alternar asistencia y aprobación
    inscripcion.asistio = not inscripcion.asistio
    inscripcion.aprobado = inscripcion.asistio
    
    if inscripcion.aprobado and not inscripcion.fecha_emision:
        inscripcion.fecha_emision = timezone.now()
        
    inscripcion.save()
    
    status_text = "APROBADO Y ASISTIÓ" if inscripcion.asistio else "PENDIENTE"
    messages.success(request, f"Estatus de {inscripcion.nombre_completo} actualizado a: {status_text}.")
    return redirect(f'/capacitaciones/admin/?curso_id={inscripcion.curso.id}')


@requiere_operador_aprobado
def aprobar_todos_capacitacion(request, curso_id):
    """
    Marca como Asistió y Aprobó a todos los participantes inscritos en el curso.
    """
    curso = get_object_or_404(CursoCapacitacion, id=curso_id)
    ahora = timezone.now()
    inscritos = InscripcionCapacitacion.objects.filter(curso=curso, asistio=False)
    total = inscritos.count()
    
    for ins in inscritos:
        ins.asistio = True
        ins.aprobado = True
        if not ins.fecha_emision:
            ins.fecha_emision = ahora
        ins.save()
        
    messages.success(request, f"¡Éxito! Se aprobaron y habilitaron las constancias de {total} participantes del curso '{curso.titulo}'.")
    return redirect(f'/capacitaciones/admin/?curso_id={curso.id}')


@requiere_operador_aprobado
def descargar_constancias_lote(request, curso_id):
    """
    Vista de lote para generar y descargar todas las constancias aprobadas en un archivo ZIP (.zip)
    o imprimirlas todas en un solo documento multi-página.
    """
    curso = get_object_or_404(CursoCapacitacion, id=curso_id)
    inscritos = InscripcionCapacitacion.objects.filter(curso=curso, asistio=True).order_by('nombre_completo')
    
    if not inscritos.exists():
        messages.warning(request, f"El taller '{curso.titulo}' aún no tiene participantes marcados con Asistencia/Aprobados. Haz clic en 'Aprobar a Todos' o marca la asistencia primero.")
        return redirect(f'/capacitaciones/admin/?curso_id={curso.id}')
        
    fecha_dt = curso.fecha_inicio or timezone.now()
    mes_nombre = MESES_ES.get(fecha_dt.month, 'Septiembre')
    fecha_pie = f"Medellin de Bravo, Ver. {fecha_dt.day} de {mes_nombre} del {fecha_dt.year}"
    
    return render(request, 'portal/capacitacion_lote_zip.html', {
        'curso': curso,
        'inscritos': inscritos,
        'fecha_pie': fecha_pie,
        'total': inscritos.count()
    })


@requiere_operador_aprobado
def editar_inscripcion_capacitacion(request, inscripcion_id):
    """
    Permite modificar los datos de un participante (nombre, CURP, empresa, teléfono, correo) desde el panel admin.
    Fuerza mayúsculas en nombre, CURP y empresa.
    """
    inscripcion = get_object_or_404(InscripcionCapacitacion, id=inscripcion_id)
    if request.method == 'POST':
        nombre = request.POST.get('nombre_completo', '').strip().upper()
        curp = request.POST.get('curp', '').strip().upper()
        empresa = request.POST.get('empresa_institucion', '').strip().upper()
        telefono = request.POST.get('telefono', '').strip()
        correo = request.POST.get('correo', '').strip()
        
        if nombre:
            inscripcion.nombre_completo = nombre
            inscripcion.curp = curp or None
            inscripcion.empresa_institucion = empresa or None
            inscripcion.telefono = telefono or None
            inscripcion.correo = correo or None
            inscripcion.save()
            messages.success(request, f"Se actualizaron con éxito los datos de '{inscripcion.nombre_completo}'.")
        else:
            messages.error(request, "El nombre completo no puede quedar en blanco.")
            
    return redirect(f'/capacitaciones/admin/?curso_id={inscripcion.curso.id}')


@requiere_operador_aprobado
def eliminar_inscripcion_capacitacion(request, inscripcion_id):
    """
    Elimina permanentemente una inscripción / registro de constancia desde el panel administrativo.
    """
    inscripcion = get_object_or_404(InscripcionCapacitacion, id=inscripcion_id)
    curso_id = inscripcion.curso.id
    nombre = inscripcion.nombre_completo
    folio = inscripcion.folio_constancia
    inscripcion.delete()
    messages.success(request, f"Se eliminó correctamente la constancia '{folio}' a nombre de {nombre}.")
    return redirect(f'/capacitaciones/admin/?curso_id={curso_id}')


@requiere_operador_aprobado
def exportar_capacitados_excel(request, curso_id):
    """
    Genera y descarga la lista oficial de inscritos y aprobados en formato Excel (.xlsx).
    """
    curso = get_object_or_404(CursoCapacitacion, id=curso_id)
    inscritos = InscripcionCapacitacion.objects.filter(curso=curso).order_by('nombre_completo')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Asistencia"
    
    # Estilos institucionales
    header_fill = PatternFill(start_color="5A123E", end_color="5A123E", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="5A123E")
    bold_font = Font(name="Calibri", size=11, bold=True)
    border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                    top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))
    
    # Título del reporte
    ws.merge_cells("A1:G1")
    ws["A1"] = f"DIRECCIÓN DE PROTECCIÓN CIVIL Y BOMBEROS — MEDELLÍN DE BRAVO"
    ws["A1"].font = title_font
    
    ws.merge_cells("A2:G2")
    ws["A2"] = f"CURSO: {curso.titulo.upper()} ({curso.duracion_horas} HRS)"
    ws["A2"].font = bold_font
    
    ws.merge_cells("A3:G3")
    ws["A3"] = f"Fecha: {curso.fecha_inicio.strftime('%d/%m/%Y')} | Sede: {curso.sede_ubicacion}"
    ws["A3"].font = Font(size=10, italic=True)
    
    headers = ["#", "Folio Constancia", "Nombre del Participante", "CURP", "Empresa / Institución", "Teléfono", "Estatus Asistencia"]
    ws.append([]) # Línea 4 en blanco
    ws.append(headers) # Línea 5
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for index, item in enumerate(inscritos, start=1):
        row = [
            index,
            item.folio_constancia,
            item.nombre_completo,
            item.curp or "-",
            item.empresa_institucion or "PARTICULAR",
            item.telefono or "-",
            "APROBADO / ASISTIÓ" if item.asistio else "PENDIENTE"
        ]
        ws.append(row)
        current_row = ws.max_row
        for col_num in range(1, 8):
            c = ws.cell(row=current_row, column=col_num)
            c.border = border
            if col_num in [1, 2, 4, 6, 7]:
                c.alignment = Alignment(horizontal="center")
                
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="Lista_Capacitacion_{curso.id}.xlsx"'
    wb.save(response)
    return response


MESES_ES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
    7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}


def imprimir_constancia_pdf(request, folio):
    """
    Vista de impresión oficial de la Constancia en formato horizontal (Landscape) con QR.
    """
    inscripcion = get_object_or_404(InscripcionCapacitacion.objects.select_related('curso'), folio_constancia=folio)
    
    if not inscripcion.asistio:
        messages.error(request, "⚠️ No se puede generar la constancia porque la persona no ha sido marcada con Asistencia/Aprobado por el Administrador.")
        return redirect(f'/capacitaciones/admin/?curso_id={inscripcion.curso.id}')
        
    if not inscripcion.fecha_emision:
        inscripcion.fecha_emision = timezone.now()
        inscripcion.save()

    # Usar la fecha del curso para la constancia, como solicitó el usuario
    fecha_dt = inscripcion.curso.fecha_inicio or inscripcion.fecha_emision or timezone.now()
    mes_nombre = MESES_ES.get(fecha_dt.month, 'Junio')
    fecha_pie = f"Medellin de Bravo, Ver. {fecha_dt.day} de {mes_nombre} del {fecha_dt.year}"
        
    return render(request, 'portal/capacitacion_imprimir_constancia.html', {
        'inscripcion': inscripcion,
        'curso': inscripcion.curso,
        'fecha_pie': fecha_pie
    })


def validar_constancia_qr(request, token):
    """
    Pantalla pública de verificación de autenticidad del documento por Código QR.
    """
    try:
        inscripcion = InscripcionCapacitacion.objects.select_related('curso').get(codigo_qr_token=token)
        valido = True
    except InscripcionCapacitacion.DoesNotExist:
        inscripcion = None
        valido = False
        
    return render(request, 'portal/capacitacion_validar_qr.html', {
        'valido': valido,
        'inscripcion': inscripcion
    })


def buscar_mis_constancias(request):
    """
    Vista pública para que los ciudadanos busquen sus constancias ingresando su nombre.
    """
    query = request.GET.get('q', '').strip()
    resultados = None
    buscado = False
    
    if query:
        buscado = True
        resultados = InscripcionCapacitacion.objects.filter(
            nombre_completo__icontains=query
        ).select_related('curso').order_by('-fecha_registro')
        
    return render(request, 'portal/capacitacion_buscar_constancias.html', {
        'query': query,
        'resultados': resultados,
        'buscado': buscado
    })


import os
import io
import qrcode
from PIL import Image, ImageDraw
from django.conf import settings

def generar_qr_registro_con_logo(request):
    """
    Genera dinámicamente la imagen PNG del Código QR con la insignia circular oficial recortada en el centro.
    Por defecto redirige al formulario de Google Forms indicado.
    """
    target_url = request.GET.get('url', 'https://forms.gle/oCSswi7Zv4zi1HDr5')
    
    qr = qrcode.QRCode(
        version=3,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=12,
        border=3,
    )
    qr.add_data(target_url)
    qr.make(fit=True)
    
    qr_img = qr.make_image(fill_color="#5A123E", back_color="white").convert('RGBA')
    qr_width, qr_height = qr_img.size
    
    logo_path = os.path.join(settings.BASE_DIR, 'portal', 'static', 'portal', 'img', 'logo_qr_pc_clean.png')
    if not os.path.exists(logo_path):
        logo_path = os.path.join(settings.BASE_DIR, 'portal', 'static', 'portal', 'img', 'logo_qr_pc.png')
        
    if os.path.exists(logo_path):
        logo = Image.open(logo_path).convert('RGBA')
        
        # Si la imagen proviene del origen original con fondo negro, aplicar máscara circular limpia
        if 'logo_qr_pc_clean.png' not in logo_path:
            w, h = logo.size
            mask = Image.new('L', (w, h), 0)
            draw_m = ImageDraw.Draw(mask)
            margin = int(w * 0.16)
            draw_m.ellipse((margin, margin, w - margin, h - margin), fill=255)
            logo.putalpha(mask)
            bbox = logo.getbbox()
            if bbox:
                logo = logo.crop(bbox)
        
        # Aumentar tamaño del logo al 30% del QR sin fondo negro
        logo_size = int(qr_width * 0.30)
        logo = logo.resize((logo_size, logo_size), Image.Resampling.LANCZOS)
        
        bg_size = logo_size + 8
        logo_bg = Image.new('RGBA', (bg_size, bg_size), (0, 0, 0, 0))
        
        draw = ImageDraw.Draw(logo_bg)
        draw.ellipse((0, 0, bg_size - 1, bg_size - 1), fill="white", outline="#E59E27", width=3)
        
        logo_bg.paste(logo, (4, 4), logo)
        
        pos_x = (qr_width - bg_size) // 2
        pos_y = (qr_height - bg_size) // 2
        qr_img.paste(logo_bg, (pos_x, pos_y), logo_bg)
        
    buffer = io.BytesIO()
    qr_img.save(buffer, format='PNG')
    return HttpResponse(buffer.getvalue(), content_type='image/png')

