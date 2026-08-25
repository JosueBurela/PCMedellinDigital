# ==============================================================================
#  📱 PARSER Y CLASIFICADOR AUTOMÁTICO DE CHATS DE WHATSAPP (REPORTES DE AGOSTO)
#  Copyright (c) 2026 Josué Jaziel Delgado Burela. Todos los derechos reservados.
# ==============================================================================

import re
import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.core.management.base import BaseCommand
from portal.models import ReporteRiesgo, Localidad

class Command(BaseCommand):
    help = 'Parsea un archivo .txt exportado de WhatsApp, extrae los reportes de Agosto y los clasifica por categoría y colonia.'

    def add_arguments(self, parser):
        parser.add_argument('--archivo', type=str, required=True, help='Ruta al archivo .txt del chat de WhatsApp exportado')
        parser.add_argument('--mes', type=int, default=8, help='Mes a filtrar (default: 8 para Agosto)')
        parser.add_argument('--year', type=int, default=2026, help='Año a filtrar (default: 2026)')
        parser.add_argument('--importar-db', action='store_true', help='Si se incluye, importa los reportes a la base de datos Django')
        parser.add_argument('--excel-output', type=str, default='Reportes_Agosto_WhatsApp.xlsx', help='Nombre del archivo Excel a generar')

    def handle(self, *args, **options):
        filepath = options['archivo']
        mes_target = options['mes']
        year_target = options['year']
        importar_db = options['importar_db']
        excel_name = options['excel_output']

        if not os.path.exists(filepath):
            self.stderr.write(f"Error: El archivo '{filepath}' no existe.")
            return

        self.stdout.write(f"Procesando chat de WhatsApp desde: {filepath}...")

        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        # Diccionario de palabras clave para clasificación de tipo de servicio
        CATEGORIAS_PATTERNS = {
            'FUEGO': [r'incendio', r'fuego', r'pastizal', r'humo', r'quema', r'basura', r'llanta', r'se quema', r'llamara'],
            'CHOQUE': [r'choque', r'accidente', r'volcadura', r'atropellad', r'derrapad', r'moto', r'auto', r'vehiculo', r'carro'],
            'GAS': [r'gas', r'fuga', r'tanque', r'cilindro', r'olor a gas', r'oler'],
            'ABEJAS': [r'abeja', r'enjambre', r'avispa', r'picadura'],
            'ARBOL': [r'arbol', r'árbol', r'rama', r'caido', r'caído', r'viento'],
            'CABLE': [r'cable', r'poste', r'transformador', r'luz', r'corto', r'chispas', r'chispa'],
            'AGUA': [r'inundac', r'agua', r'rio', r'río', r'anegad', r'inundado', r'drenaje', r'canal'],
            'OTRO': []
        }

        # Localidades conocidas de Medellín de Bravo
        LOCALIDADES_MEDELLIN = [
            'PUENTE MORENO', 'ARBOLEDAS SAN RAMÓN', 'LAGOS DE PUENTE MORENO', 'EL TEJAR',
            'MEDELLÍN', 'PLAYA DE VACAS', 'PASO DEL TORO', 'LOS ROBLES', 'DOS BOCAS',
            'RANCHO DEL PADRE', 'PASO COLORADO', 'LA JOYA', 'PASEO CAMPESTRE', 'HERÓN PROAL',
            'MARCOS VÉLEZ', 'EMILIANO ZAPATA', 'CLARA CÓRDOBA'
        ]

        # Regex para timestamps típicos de WhatsApp:
        # 1) [dd/mm/aaaa, hh:mm:ss]
        # 2) dd/mm/aaaa, hh:mm -
        pattern_ts = re.compile(r'^(?:\[)?(\d{1,2}/\d{1,2}/\d{2,4})[,\s]+(\d{1,2}:\d{2}(?::\d{2})?(?:\s*[ap]\.?\s*m\.?)?)(?:\])?\s*[-–]?\s*(.*?):\s*(.*)$', re.IGNORECASE)

        reportes_extraidos = []
        msg_buffer = None

        for line in lines:
            line_str = line.strip()
            match = pattern_ts.match(line_str)

            if match:
                if msg_buffer:
                    reportes_extraidos.append(msg_buffer)
                    msg_buffer = None

                fecha_raw, hora_raw, remitente, texto = match.groups()

                # Parsear fecha
                try:
                    parts = fecha_raw.split('/')
                    d = int(parts[0])
                    m = int(parts[1])
                    y = int(parts[2])
                    if y < 100: y += 2000
                    dt = datetime.date(y, m, d)
                except Exception:
                    continue

                # Filtrar solo por el mes y año objetivo (ej. Agosto)
                if dt.month == mes_target and dt.year == year_target:
                    msg_buffer = {
                        'fecha': dt,
                        'fecha_str': dt.strftime('%d/%m/%Y'),
                        'hora': hora_raw.strip(),
                        'remitente': remitente.strip(),
                        'texto': texto.strip()
                    }
            else:
                # Líneas multilínea del mismo mensaje
                if msg_buffer:
                    msg_buffer['texto'] += f" {line_str}"

        if msg_buffer:
            reportes_extraidos.append(msg_buffer)

        # Ignorar mensajes de sistema (ej. "Los mensajes están cifrados", "cambió el icono", etc.)
        reportes_filtrados = []
        for r in reportes_extraidos:
            txt_lower = r['texto'].lower()
            if any(sys_kw in txt_lower for sys_kw in ['cambió', 'añadió', 'eliminó', 'salio', 'salió', 'cifrado', 'código de seguridad']):
                continue
            if len(r['texto']) < 5:
                continue

            # Clasificar Categoría
            tipo_detectado = 'OTRO'
            for cat, kw_list in CATEGORIAS_PATTERNS.items():
                if any(re.search(kw, txt_lower) for kw in kw_list):
                    tipo_detectado = cat
                    break
            r['tipo_servicio'] = tipo_detectado

            # Detectar Localidad / Colonia
            loc_detectada = 'MEDELLÍN'
            for loc in LOCALIDADES_MEDELLIN:
                if loc.lower() in txt_lower:
                    loc_detectada = loc
                    break
            r['localidad'] = loc_detectada

            reportes_filtrados.append(r)

        self.stdout.write(self.style.SUCCESS(f"¡Extracción completada! Se encontraron {len(reportes_filtrados)} reportes del mes de Agosto."))

        # 1. EXPORTAR A EXCEL
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reportes Agosto 2026"

        header_fill = PatternFill(start_color="5A123E", end_color="5A123E", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="5A123E")
        border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

        ws.merge_cells("A1:F1")
        ws["A1"] = f"PROTECCIÓN CIVIL DE MEDELLÍN DE BRAVO — BITÁCORA DE REPORTES (AGOSTO {year_target})"
        ws["A1"].font = title_font

        headers = ["#", "Fecha", "Hora", "Remitente / Ciudadano", "Categoría / Incidente", "Colonia / Localidad", "Texto del Reporte"]
        ws.append([]) # Fila 2 en blanco
        ws.append(headers) # Fila 3

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, item in enumerate(reportes_filtrados, 1):
            row = [
                idx,
                item['fecha_str'],
                item['hora'],
                item['remitente'],
                item['tipo_servicio'],
                item['localidad'],
                item['texto']
            ]
            ws.append(row)
            r_num = ws.max_row
            for c_i in range(1, 8):
                c = ws.cell(row=r_num, column=c_i)
                c.border = border
                if c_i in [1, 2, 3, 5, 6]:
                    c.alignment = Alignment(horizontal="center")

        wb.save(excel_name)
        self.stdout.write(self.style.SUCCESS(f"✔ Archivo Excel generado con éxito: {excel_name}"))

        # 2. IMPORTAR A BASE DE DATOS DJANGO (SI SE INDICA EL FLAG)
        if importar_db:
            creados = 0
            for item in reportes_filtrados:
                # Generar folio único para el reporte
                count = ReporteRiesgo.objects.count() + 1
                folio = f"REP-{year_target}-WA-{count:04d}"

                ReporteRiesgo.objects.create(
                    numero_reporte=folio,
                    nombre_ciudadano=item['remitente'][:150],
                    telefono_ciudadano='2290000000',
                    tipo_servicio=item['tipo_servicio'],
                    descripcion=item['texto'],
                    localidad=item['localidad'],
                    ubicacion_direccion=f"Reportado vía WhatsApp en {item['localidad']}",
                    estatus='RESUELTO'
                )
                creados += 1
            self.stdout.write(self.style.SUCCESS(f"✔ Se importaron {creados} reportes a la base de datos de Protección Civil."))
