# ==============================================================================
#  📱 EXTRACTOR DE REPORTES DE WHATSAPP DIRECTO DE EVOLUTION API LOCAL
#  Copyright (c) 2026 Josué Jaziel Delgado Burela. Todos los derechos reservados.
# ==============================================================================

import os
import re
import io
import json
import datetime
import urllib.request
import urllib.error
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.core.management.base import BaseCommand
from portal.models import ReporteRiesgo

class Command(BaseCommand):
    help = 'Conecta con la API local de Evolution API (http://localhost:8080), extrae todos los mensajes del grupo de WhatsApp y genera el Excel de reportes de Agosto.'

    def add_arguments(self, parser):
        parser.add_argument('--url', type=str, default='http://localhost:8080', help='URL de Evolution API local')
        parser.add_argument('--apikey', type=str, default='MedellinPCSecretToken2026', help='API Key de Evolution API local')
        parser.add_argument('--instancia', type=str, default='PCMedellin', help='Nombre de la instancia de WhatsApp local')
        parser.add_argument('--group-jid', type=str, default='', help='JID del grupo de WhatsApp (si no se indica, listará todos los grupos disponibles)')
        parser.add_argument('--mes', type=int, default=8, help='Mes a filtrar (default: 8 para Agosto)')
        parser.add_argument('--year', type=int, default=2026, help='Año a filtrar (default: 2026)')
        parser.add_argument('--excel-output', type=str, default='Reportes_Agosto_WhatsApp_Local.xlsx', help='Nombre del archivo Excel a generar')
        parser.add_argument('--importar-db', action='store_true', help='Si se incluye, importa los reportes a la base de datos local')

    def handle(self, *args, **options):
        api_url = options['url'].rstrip('/')
        api_key = options['apikey']
        instancia = options['instancia']
        group_jid = options['group_jid']
        mes_target = options['mes']
        year_target = options['year']
        excel_name = options['excel_output']
        importar_db = options['importar_db']

        self.stdout.write(self.style.SUCCESS(f"Conectando con Evolution API Local en {api_url} (Instancia: {instancia})..."))

        # Si no nos pasaron el group_jid, consultamos los grupos disponibles en la API local
        if not group_jid:
            try:
                req_groups = urllib.request.Request(f"{api_url}/group/fetchAllGroups/{instancia}?getParticipants=false")
                req_groups.add_header('apikey', api_key)
                with urllib.request.urlopen(req_groups) as resp:
                    groups_data = json.loads(resp.read().decode('utf-8'))
                    
                if not groups_data:
                    self.stderr.write("No se encontraron grupos asociados en la cuenta de WhatsApp conectada.")
                    return
                
                self.stdout.write(self.style.SUCCESS(f"\n--- GRUPOS ENCONTRADOS EN TU WHATSAPP LOCAL ({len(groups_data)}) ---"))
                for g in groups_data:
                    self.stdout.write(f" -> JID: {g.get('id')} | Grupo: {g.get('subject')}")
                
                # Seleccionar por defecto el primer grupo o el grupo de alertas
                group_jid = groups_data[0].get('id')
                self.stdout.write(f"\nProcesando automáticamente el grupo: '{groups_data[0].get('subject')}' ({group_jid})...\n")

            except Exception as e:
                self.stderr.write(f"Error conectando a Evolution API local: {e}")
                return

        # Consultar los mensajes del grupo
        payload = {
            'where': {
                'key': {
                    'remoteJid': group_jid
                }
            },
            'limit': 1000
        }

        try:
            req_msg = urllib.request.Request(
                f"{api_url}/chat/findMessages/{instancia}",
                data=json.dumps(payload).encode('utf-8'),
                headers={
                    'apikey': api_key,
                    'Content-Type': 'application/json'
                }
            )
            with urllib.request.urlopen(req_msg) as resp:
                res_data = json.loads(resp.read().decode('utf-8'))
                records = res_data.get('messages', {}).get('records', [])
        except Exception as e:
            self.stderr.write(f"Error consultando mensajes del grupo: {e}")
            return

        self.stdout.write(f"Se recuperaron {len(records)} mensajes del grupo.")

        # Diccionario de patrones para clasificación
        CATEGORIAS_PATTERNS = {
            'FUEGO': [r'incendio', r'fuego', r'pastizal', r'humo', r'quema', r'basura', r'llanta', r'se quema'],
            'CHOQUE': [r'choque', r'accidente', r'volcadura', r'atropellad', r'derrapad', r'moto', r'auto', r'vehiculo', r'carro'],
            'GAS': [r'gas', r'fuga', r'tanque', r'cilindro', r'olor a gas', r'oler'],
            'ABEJAS': [r'abeja', r'enjambre', r'avispa', r'picadura'],
            'ARBOL': [r'arbol', r'árbol', r'rama', r'caido', r'caído', r'viento'],
            'CABLE': [r'cable', r'poste', r'transformador', r'luz', r'corto', r'chispas'],
            'AGUA': [r'inundac', r'agua', r'rio', r'río', r'anegad', r'inundado', r'drenaje', r'canal'],
            'OTRO': []
        }

        LOCALIDADES_MEDELLIN = [
            'PUENTE MORENO', 'ARBOLEDAS SAN RAMÓN', 'LAGOS DE PUENTE MORENO', 'EL TEJAR',
            'MEDELLÍN', 'PLAYA DE VACAS', 'PASO DEL TORO', 'LOS ROBLES', 'DOS BOCAS',
            'RANCHO DEL PADRE', 'PASO COLORADO', 'LA JOYA', 'PASEO CAMPESTRE', 'HERÓN PROAL'
        ]

        reportes_filtrados = []

        for item in records:
            # Obtener timestamp del mensaje
            ts = item.get('messageTimestamp')
            if not ts:
                continue

            dt = datetime.datetime.fromtimestamp(int(ts))

            # Filtrar por mes y año
            if dt.month == mes_target and dt.year == year_target:
                msg_obj = item.get('message', {})
                txt = msg_obj.get('conversation') or msg_obj.get('extendedTextMessage', {}).get('text') or ''
                txt_clean = txt.strip()

                if len(txt_clean) < 5:
                    continue

                txt_lower = txt_clean.lower()
                if any(sys_kw in txt_lower for sys_kw in ['cambió', 'añadió', 'eliminó', 'salio', 'salió', 'cifrado', 'código de seguridad']):
                    continue

                key = item.get('key', {})
                remitente = item.get('pushName') or key.get('participant', 'Usuario WhatsApp')

                # Clasificar Incidente
                tipo_detectado = 'OTRO'
                for cat, kw_list in CATEGORIAS_PATTERNS.items():
                    if any(re.search(kw, txt_lower) for kw in kw_list):
                        tipo_detectado = cat
                        break

                # Detectar Colonia
                loc_detectada = 'MEDELLÍN'
                for loc in LOCALIDADES_MEDELLIN:
                    if loc.lower() in txt_lower:
                        loc_detectada = loc
                        break

                reportes_filtrados.append({
                    'fecha_str': dt.strftime('%d/%m/%Y'),
                    'hora': dt.strftime('%H:%M:%S'),
                    'remitente': remitente,
                    'tipo_servicio': tipo_detectado,
                    'localidad': loc_detectada,
                    'texto': txt_clean
                })

        self.stdout.write(self.style.SUCCESS(f"\n¡Se filtraron {len(reportes_filtrados)} reportes del mes de Agosto ({mes_target}/{year_target})!"))

        # Generar Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Reportes Mes {mes_target}"

        header_fill = PatternFill(start_color="5A123E", end_color="5A123E", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="5A123E")
        border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

        ws.merge_cells("A1:G1")
        ws["A1"] = f"DIRECCIÓN DE PROTECCIÓN CIVIL — REPORTES EXTRAÍDOS DE EVOLUTION API LOCAL ({mes_target}/{year_target})"
        ws["A1"].font = title_font

        headers = ["#", "Fecha", "Hora", "Remitente", "Categoría Incidente", "Colonia / Localidad", "Descripción del Reporte"]
        ws.append([])
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, item in enumerate(reportes_filtrados, 1):
            row = [idx, item['fecha_str'], item['hora'], item['remitente'], item['tipo_servicio'], item['localidad'], item['texto']]
            ws.append(row)
            r_num = ws.max_row
            for c_i in range(1, 8):
                c = ws.cell(row=r_num, column=c_i)
                c.border = border
                if c_i in [1, 2, 3, 5, 6]:
                    c.alignment = Alignment(horizontal="center")

        wb.save(excel_name)
        self.stdout.write(self.style.SUCCESS(f"=== Archivo Excel local generado con exito: {excel_name} ==="))

        if importar_db:
            creados = 0
            for item in reportes_filtrados:
                count = ReporteRiesgo.objects.count() + 1
                folio = f"REP-{year_target}-WA-{count:04d}"
                ReporteRiesgo.objects.create(
                    numero_reporte=folio,
                    nombre_ciudadano=item['remitente'][:150],
                    telefono_ciudadano='2290000000',
                    tipo_servicio=item['tipo_servicio'],
                    descripcion=item['texto'],
                    localidad=item['localidad'],
                    ubicacion_direccion=f"Reportado vía WhatsApp en {item['localidad']}",
                    estatus='RESUELTO'
                )
                creados += 1
            self.stdout.write(self.style.SUCCESS(f"=== Se importaron {creados} reportes a tu base de datos Django local. ==="))
