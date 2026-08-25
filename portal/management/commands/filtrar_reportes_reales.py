# ==============================================================================
#  🧠 MODELO INTELIGENTE DE FILTRADO Y ESTRUCTURACIÓN DE REPORTES DE EMERGENCIA
#  Copyright (c) 2026 Josué Jaziel Delgado Burela. Todos los derechos reservados.
# ==============================================================================

import openpyxl
import re
import os
import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.core.management.base import BaseCommand

class Command(BaseCommand):
    help = 'Analiza y filtra los reportes reales de emergencia excluyendo el ruido de conversación de WhatsApp del mes de Agosto.'

    def add_arguments(self, parser):
        parser.add_argument('--input', type=str, default='Reportes_Agosto_Completo_Proteccion_Civil.xlsx', help='Archivo de origen con los 2300 mensajes')
        parser.add_argument('--output', type=str, default='Reportes_Agosto_Estructurado_Oficial.xlsx', help='Nombre del Excel final estructurado')

    def handle(self, *args, **options):
        input_file = options['input']
        output_file = options['output']

        if not os.path.exists(input_file):
            self.stderr.write(f"Error: El archivo '{input_file}' no existe.")
            return

        self.stdout.write(self.style.SUCCESS(f"Analizando y estructurando reportes desde '{input_file}'..."))

        wb_in = openpyxl.load_workbook(input_file)
        ws_in = wb_in.active

        rows = list(ws_in.iter_rows(values_only=True))[3:] # Ignorar encabezados

        PATRONES_RUIDO = [
            r'^\s*enterado\b', r'^\s*recibido\b', r'^\s*ok\b', r'^\s*gracias\b', r'^\s*de acuerdo\b',
            r'^\s*en base\b', r'^\s*unidad \d+ en base\b', r'^\s*de retorno a base\b', r'^\s*sin novedad\b',
            r'^\s*entendido\b', r'^\s*copiado\b', r'^\s*enterados\b', r'^\s*enterado jefe\b', r'^\s*enterado gracias\b',
            r'^\s*enterado en base\b', r'^\s*enterado con precaución\b'
        ]

        PATRONES_EMERGENCIA = [
            r'report', r'fuga', r'incendio', r'fuego', r'choque', r'accidente', r'volcadura', r'atropellad',
            r'abeja', r'enjambre', r'avispa', r'arbol', r'árbol', r'rama', r'cable', r'poste', r'corto',
            r'inundac', r'agua', r'derram', r'tanque', r'cilindro', r'lesionad', r'herid', r'auxilio',
            r'apoyo', r'colonia', r'calle', r'av\.', r'avenida', r'fracc', r'ambulancia', r'traslado', r'paciente',
            r'hospital', r'paramedico', r'operador', r'atendi'
        ]

        LOCALIDADES_MEDELLIN = [
            'PUENTE MORENO', 'ARBOLEDAS SAN RAMÓN', 'LAGOS DE PUENTE MORENO', 'EL TEJAR',
            'MEDELLÍN', 'PLAYA DE VACAS', 'PASO DEL TORO', 'LOS ROBLES', 'DOS BOCAS',
            'RANCHO DEL PADRE', 'PASO COLORADO', 'LA JOYA', 'PASEO CAMPESTRE', 'HERÓN PROAL'
        ]

        reportes_reales = []
        novedades_base = []

        conteo_categorias = {}
        conteo_localidades = {}

        for r in rows:
            if not r or len(r) < 7: continue
            idx, fecha, hora, remitente, cat, loc, texto = r
            if not texto: continue

            txt_str = str(texto).strip()
            txt_lower = txt_str.lower()

            # Descartar ruidos de chat
            is_ruido = any(re.search(p, txt_lower) for p in PATRONES_RUIDO) and len(txt_str) < 60
            has_emergencia = any(re.search(p, txt_lower) for p in PATRONES_EMERGENCIA) or len(txt_str) > 70

            if is_ruido or not has_emergencia:
                novedades_base.append(r)
            else:
                # Recalcular categoría precisa
                cat_final = cat
                if cat == 'OTRO':
                    if any(re.search(p, txt_lower) for p in [r'ambulancia', r'traslado', r'paciente', r'hospital', r'herid', r'paramedico']):
                        cat_final = 'ATENCIÓN MÉDICA / TRASLADO'
                    elif any(re.search(p, txt_lower) for p in [r'fuego', r'incendio', r'pastizal', r'quema']):
                        cat_final = 'INCENDIO / QUEMA'
                    elif any(re.search(p, txt_lower) for p in [r'choque', r'accidente', r'volcadura', r'atropellad']):
                        cat_final = 'ACCIDENTE VIAL'
                    elif any(re.search(p, txt_lower) for p in [r'gas', r'fuga', r'tanque']):
                        cat_final = 'FUGA DE GAS'
                    elif any(re.search(p, txt_lower) for p in [r'abeja', r'enjambre', r'avispa']):
                        cat_final = 'ENJAMBRE DE ABEJAS'
                    elif any(re.search(p, txt_lower) for p in [r'cable', r'poste', r'luz', r'corto']):
                        cat_final = 'CABLE / POSTE EXPUESTO'
                    elif any(re.search(p, txt_lower) for p in [r'arbol', r'árbol', r'rama']):
                        cat_final = 'ÁRBOL CAÍDO / OBSTRUCCIÓN'
                    else:
                        cat_final = 'INCIDENTE OPERATIVO PC'

                # Conteo estadístico
                conteo_categorias[cat_final] = conteo_categorias.get(cat_final, 0) + 1
                conteo_localidades[loc] = conteo_localidades.get(loc, 0) + 1

                reportes_reales.append((
                    len(reportes_reales) + 1,
                    fecha,
                    hora,
                    remitente,
                    cat_final,
                    loc,
                    txt_str
                ))

        self.stdout.write(self.style.SUCCESS(f"=== Filtrado completado: {len(reportes_reales)} Reportes Reales de Emergencia identificados de 2,300 mensajes. ==="))

        # ----------------------------------------------------------------------
        # CREACIÓN DEL LIBRO EXCEL ESTRUCTURADO Y EJECUTIVO
        # ----------------------------------------------------------------------
        wb_out = openpyxl.Workbook()

        header_fill = PatternFill(start_color="5A123E", end_color="5A123E", fill_type="solid")
        gold_fill = PatternFill(start_color="E59E27", end_color="E59E27", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="5A123E")
        sub_font = Font(name="Calibri", size=12, bold=True, color="1E293B")
        border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

        # PESTAÑA 1: REPORTES REALES DE EMERGENCIA
        ws1 = wb_out.active
        ws1.title = "Reportes Reales Emergencia"

        ws1.merge_cells("A1:G1")
        ws1["A1"] = "DIRECCIÓN DE PROTECCIÓN CIVIL DE MEDELLÍN — BITÁCORA OFICIAL DE REPORTES REALES DE EMERGENCIA (AGOSTO 2026)"
        ws1["A1"].font = title_font

        headers1 = ["Folio #", "Fecha", "Hora", "Reportante / Personal", "Tipo de Emergencia", "Colonia / Localidad", "Detalle del Reporte de Emergencia"]
        ws1.append([])
        ws1.append(headers1)

        for col_num, header in enumerate(headers1, 1):
            cell = ws1.cell(row=3, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in reportes_reales:
            ws1.append(list(r))
            r_num = ws1.max_row
            for c_i in range(1, 8):
                c = ws1.cell(row=r_num, column=c_i)
                c.border = border
                if c_i in [1, 2, 3, 5, 6]:
                    c.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    c.alignment = Alignment(vertical="top", wrap_text=True)

        ws1.column_dimensions['A'].width = 10
        ws1.column_dimensions['B'].width = 14
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 24
        ws1.column_dimensions['E'].width = 28
        ws1.column_dimensions['F'].width = 24
        ws1.column_dimensions['G'].width = 75

        # PESTAÑA 2: ESTADÍSTICAS Y RESUMEN EJECUTIVO DEL MES
        ws2 = wb_out.create_sheet(title="Estadisticas y Resumen")

        ws2.merge_cells("A1:E1")
        ws2["A1"] = "RESUMEN EJECUTIVO Y ESTADÍSTICAS DE EMERGENCIAS (AGOSTO 2026)"
        ws2["A1"].font = title_font

        ws2["A3"] = "Total Mensajes Analizados:"
        ws2["B3"] = len(rows)
        ws2["A4"] = "Reportes Reales Atendidos:"
        ws2["B4"] = len(reportes_reales)
        ws2["A3"].font = sub_font
        ws2["A4"].font = sub_font
        ws2["B3"].font = sub_font
        ws2["B4"].font = sub_font

        # Tabla Categorías
        ws2["A6"] = "DESGLOSE POR TIPO DE EMERGENCIA ATENDIDA"
        ws2["A6"].font = sub_font
        ws2.append(["Categoría / Tipo de Servicio", "Cantidad Atendida", "Porcentaje (%)"])
        row_cat_start = ws2.max_row
        for cat_name, cnt in sorted(conteo_categorias.items(), key=lambda x: x[1], reverse=True):
            pct = round((cnt / len(reportes_reales)) * 100, 1)
            ws2.append([cat_name, cnt, f"{pct}%"])

        for col in range(1, 4):
            c = ws2.cell(row=row_cat_start, column=col)
            c.fill = gold_fill
            c.font = Font(bold=True)

        # Tabla Localidades
        ws2["D6"] = "EMERGENCIAS POR ZONA / LOCALIDAD"
        ws2["D6"].font = sub_font
        ws2.cell(row=7, column=4, value="Colonia / Localidad").fill = gold_fill
        ws2.cell(row=7, column=4).font = Font(bold=True)
        ws2.cell(row=7, column=5, value="Incidentes").fill = gold_fill
        ws2.cell(row=7, column=5).font = Font(bold=True)

        curr_r = 8
        for loc_name, cnt in sorted(conteo_localidades.items(), key=lambda x: x[1], reverse=True):
            ws2.cell(row=curr_r, column=4, value=loc_name)
            ws2.cell(row=curr_r, column=5, value=cnt)
            curr_r += 1

        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 30
        ws2.column_dimensions['E'].width = 18

        # PESTAÑA 3: NOVEDADES Y RUIDO DE BASE
        ws3 = wb_out.create_sheet(title="Bitacora Base y Mensajes")
        ws3.merge_cells("A1:G1")
        ws3["A1"] = "BITÁCORA DE NOVEDADES OPERATIVAS Y MENSAJES DE BASE (AGOSTO 2026)"
        ws3["A1"].font = title_font

        headers3 = ["#", "Fecha", "Hora", "Remitente", "Categoría", "Localidad", "Mensaje"]
        ws3.append([])
        ws3.append(headers3)

        for col_num, header in enumerate(headers3, 1):
            cell = ws3.cell(row=3, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        for idx, r in enumerate(novedades_base, 1):
            ws3.append([idx, r[1], r[2], r[3], r[4], r[5], r[6]])

        wb_out.save(output_file)
        self.stdout.write(self.style.SUCCESS(f"=== Reporte Ejecutivo estructurado guardado exitosamente en: '{output_file}' ==="))
